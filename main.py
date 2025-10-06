import locale
import re
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import validate_call, ConfigDict

from models.work_day import WorkDay

START_APPRENTICESHIP = datetime(year=2024, month=8, day=1)
TEMPLATE_PATH = Path("assets/template")
RESULT_PATH = Path("result.xlsx")
INPUT_PATH = Path("assets/input")
HOUR_KEY = "h"
DATE_KEY = "datum"
YEAR_KEY = "jahr"
MONTH_KEY = "monat"
TEXT_KEY = "kommentar"
WORKSHEET_KEY = "buchungen"
VACATION_ALIAS_KEY = "urlaub"
SICK_DAY_ALIAS_KEY = "krank tage"
SCHOOL_ALIAS_KEY = "zpe azubi ext"

_WORKBOOK: Workbook | None = None
_WORKSHEET_INDEX = 1
_ACTIVE_WORKSHEET_END_DATE: datetime | None = None

def alias_to_location(alias: str) -> str | None:
    normalized_alias = alias.lower().strip()
    if normalized_alias == VACATION_ALIAS_KEY or normalized_alias == SICK_DAY_ALIAS_KEY:
        return None
    elif normalized_alias == SCHOOL_ALIAS_KEY:
        return "HEINZ NIXDORF BERUFSKOLLEG"
    else:
        return "SOPTIM AG"

@validate_call
def get_workdays_from_workbook(workbook_path: Path) -> list[WorkDay]:
    workbook = load_workbook(filename=workbook_path)

    data_sheet: Worksheet | None = None
    for sheet in workbook.worksheets:
        if sheet.title.lower() == WORKSHEET_KEY:
            data_sheet = sheet
            break
    else:
        raise ValueError(f"Worksheet named '{WORKSHEET_KEY}' not found in {TEMPLATE_PATH}")

    start_row_date, key_column_positions_date = get_key_positions_in_worksheet(data_sheet, YEAR_KEY, MONTH_KEY)
    year = int(data_sheet.cell(row=start_row_date + 1, column=key_column_positions_date[YEAR_KEY]).value)
    month = int(data_sheet.cell(row=start_row_date + 1, column=key_column_positions_date[MONTH_KEY]).value)

    start_row_data, key_column_positions = get_key_positions_in_worksheet(data_sheet, HOUR_KEY, TEXT_KEY, DATE_KEY)
    work_days: list[WorkDay] = list()
    while start_row_data <= data_sheet.max_row:
        start_row_data += 1
        if data_sheet.cell(row=start_row_data, column=key_column_positions[HOUR_KEY]).data_type != 'n':
            break

        hours = data_sheet.cell(row=start_row_data, column=key_column_positions[HOUR_KEY]).value
        alias = data_sheet.cell(row=start_row_data,column=key_column_positions[HOUR_KEY] + 1).value
        normalized_alias = alias.lower().strip()
        text = data_sheet.cell(row=start_row_data, column=key_column_positions[TEXT_KEY]).value
        if text is None:
            if normalized_alias == VACATION_ALIAS_KEY:
                text = "Urlaub"
            elif normalized_alias == SICK_DAY_ALIAS_KEY:
                text = "Krank"
            else:
                text = alias

        date_string = data_sheet.cell(row=start_row_data, column=key_column_positions[DATE_KEY]).value
        # Extract digits from date_string to form the day
        day = int(''.join(re.findall(r'\d', date_string)))
        date = datetime(year=year, month=month, day=day)

        work_days.append(WorkDay(date=date, hours_worked=hours, text=text, location=alias_to_location(normalized_alias)))
    else:
        raise ValueError("No terminating condition found for workday entries")
    return work_days


@validate_call(config=ConfigDict(arbitrary_types_allowed=True))
def get_key_positions_in_worksheet(worksheet: Worksheet, *keys: str) -> tuple[int, dict[str, int]]:
    key_column_positions = dict()

    key_set = set()
    for key in keys:
        key_set.add(key.lower().strip())

    for row_index, row in enumerate(worksheet.iter_rows()):
        # normalizing the values by removing None, lowercasing and whitespace
        row_set = set()
        for item in row:
            if isinstance(item.value, str):
                row_set.add(item.value.strip().lower())
        # check if all values in key_set are in row_set
        if not key_set.issubset(row_set):
            continue
        for cell in row:
            if cell.value is None or not isinstance(cell.value, str):
                continue
            normalized_cell_value = cell.value.lower().strip()
            if normalized_cell_value in key_set:
                key_column_positions[normalized_cell_value] = cell.column
        return row_index + 1, key_column_positions
    raise ValueError(f"Keys {key_set} not found in any row of worksheet")

@validate_call(config=ConfigDict(arbitrary_types_allowed=True))
def insert_workday_into_workbook(work_day: WorkDay) -> None:
    merged_cell_ranges = _WORKBOOK.active.merged_cells.ranges
    start_row, column_position = get_key_positions_in_worksheet(_WORKBOOK.active, work_day.normalized_day_name)
    day_cell = _WORKBOOK.active.cell(start_row, column_position[work_day.normalized_day_name])

    # Find the first empty cell in the merged range for the day
    text_cell: Cell | None = None
    last_row_merged_cells_day: int | None = None
    for merged_cells in merged_cell_ranges:
        if day_cell.coordinate not in merged_cells:
            continue

        for index_row in range(merged_cells.min_row, merged_cells.max_row + 1):
            cell = _WORKBOOK.active.cell(index_row, column_position[work_day.normalized_day_name] + 1)
            if not cell.value:
                text_cell = cell
                text_cell.value = work_day.text
                last_row_merged_cells_day = merged_cells.max_row
                break
    if text_cell is None:
        raise ValueError(f"No empty cell found for day {work_day.normalized_day_name}")

    for merged_cells in merged_cell_ranges:
        if text_cell.coordinate in merged_cells:

            _WORKBOOK.active.cell(text_cell.row, merged_cells.max_col + 1).value = work_day.hours_worked
            _WORKBOOK.active.cell(last_row_merged_cells_day, merged_cells.max_col + 3).value = work_day.location
            break
    else:
        raise ValueError(f"Junge wie ist das überhaupt möglich")

def duplicate_and_activate_new_worksheet() -> None:
    global _WORKSHEET_INDEX
    global _ACTIVE_WORKSHEET_END_DATE

    _WORKBOOK.copy_worksheet(_WORKBOOK.worksheets[0])
    _WORKBOOK.active = _WORKBOOK.worksheets[-1]
    # Insert the worksheet index into the first row
    first_row = next(_WORKBOOK.active.iter_rows())
    for cell in first_row:
        if cell.value and isinstance(cell.value, int):
            cell.value = _WORKSHEET_INDEX

            start_weekday = START_APPRENTICESHIP.weekday()
            # Move back to the Monday of the same week
            start_monday = START_APPRENTICESHIP - timedelta(days=start_weekday)
            _ACTIVE_WORKSHEET_END_DATE = start_monday + timedelta(days=4, weeks=_WORKSHEET_INDEX - 1)

            _WORKSHEET_INDEX += 1
            break
    else:
        raise ValueError("No cell with integer value found in the first row of the template worksheet")

    date_format = "%d.%m.%y"
    _WORKBOOK.active.title = f"{(_ACTIVE_WORKSHEET_END_DATE - timedelta(days=4)).strftime(date_format)} - {_ACTIVE_WORKSHEET_END_DATE.strftime(date_format)}"

def main():
    global _WORKBOOK
    locale.setlocale(locale.LC_TIME, "")

    first_file_in_template_dir = next(TEMPLATE_PATH.iterdir())
    if first_file_in_template_dir.exists():
        _WORKBOOK = load_workbook(filename=first_file_in_template_dir)
        _WORKBOOK.template = False
        duplicate_and_activate_new_worksheet()
    else:
        raise FileNotFoundError(f"No template file found at {TEMPLATE_PATH}")

    for file_path in INPUT_PATH.iterdir():
        if file_path.name.startswith("."):
            continue
        workdays = get_workdays_from_workbook(file_path)
        for workday in workdays:
            while workday.date > _ACTIVE_WORKSHEET_END_DATE:
                duplicate_and_activate_new_worksheet()
            insert_workday_into_workbook(workday)

    _WORKBOOK.remove(_WORKBOOK.worksheets[0])
    _WORKBOOK.save(RESULT_PATH)

if __name__ == "__main__":
    main()